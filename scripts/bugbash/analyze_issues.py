#!/usr/bin/env python3
"""Analyze GitHub issues from an Excel spreadsheet.

Takes an Excel file (.xlsx) where each sheet represents an area.
Each sheet should have a column containing GitHub issue URLs.

Reports:
1. Total number of closed issues
2. Number of closed issues per area (sheet)
3. Number of issues closed by each user
4. High priority / UBN issues closed by each user
5. Number of PRs reviewed per user across linked PRs
6. Number of landed issues per assignee (from sheet status column)
7. Number of landed issues per area (from sheet status column)

Requires: gh CLI (authenticated), openpyxl
"""

import argparse
import json
import re
import subprocess
import sys
from collections import defaultdict


try:
    import openpyxl
except ImportError:
    print("Error: openpyxl is required. Install with: pip install openpyxl")
    sys.exit(1)


GITHUB_ISSUE_RE = re.compile(r"https?://github\.com/([^/]+)/([^/]+)/issues/(\d+)")
BOTS = {"pytorchmergebot", "facebook-github-bot", "pytorch-bot[bot]"}
HIGH_PRIORITY_LABELS = {"high priority", "pt2: ubn"}


def parse_issue_url(url: str) -> tuple[str, str, int] | None:
    """Extract (owner, repo, issue_number) from a GitHub issue URL."""
    m = GITHUB_ISSUE_RE.search(url)
    if m:
        return m.group(1), m.group(2), int(m.group(3))
    return None


def gh_json(args: list[str]) -> dict | list:
    """Run a gh command with --json output and return parsed JSON."""
    result = subprocess.run(
        ["gh"] + args,
        capture_output=True,
        text=True,
    )
    if result.returncode != 0:
        raise RuntimeError(f"gh command failed: {result.stderr.strip()}")
    return json.loads(result.stdout)


def gh_json_paginated(args: list[str]) -> list:
    """Run a gh api command with --paginate and return merged JSON arrays."""
    result = subprocess.run(
        ["gh"] + args,
        capture_output=True,
        text=True,
    )
    if result.returncode != 0:
        raise RuntimeError(f"gh command failed: {result.stderr.strip()}")
    decoder = json.JSONDecoder()
    output = result.stdout.strip()
    if not output:
        return []
    items: list = []
    pos = 0
    while pos < len(output):
        obj, end = decoder.raw_decode(output, pos)
        if isinstance(obj, list):
            items.extend(obj)
        else:
            items.append(obj)
        pos = end
        while pos < len(output) and output[pos] in " \t\n\r":
            pos += 1
    return items


def get_issue_info(owner: str, repo: str, number: int) -> dict:
    """Fetch issue state, closing PRs, and title."""
    return gh_json(
        [
            "issue",
            "view",
            str(number),
            "--repo",
            f"{owner}/{repo}",
            "--json",
            "state,author,closedByPullRequestsReferences,title,labels",
        ]
    )


def get_issue_timeline(owner: str, repo: str, number: int) -> list[dict]:
    """Fetch the full paginated timeline for an issue."""
    try:
        return gh_json_paginated(
            [
                "api",
                f"repos/{owner}/{repo}/issues/{number}/timeline",
                "--paginate",
            ]
        )
    except RuntimeError:
        return []


def extract_pr_numbers_from_timeline(events: list[dict]) -> set[int]:
    """Extract all cross-referenced PR numbers from timeline events."""
    pr_numbers = set()
    for event in events:
        if not isinstance(event, dict):
            continue
        source = event.get("source", {})
        if not source:
            continue
        issue = source.get("issue", {})
        pr = issue.get("pull_request", {})
        if pr and pr.get("html_url"):
            m = re.search(r"/pull/(\d+)", pr["html_url"])
            if m:
                pr_numbers.add(int(m.group(1)))
    return pr_numbers


def find_closer_from_timeline(events: list[dict], owner: str, repo: str) -> str | None:
    """Find who closed an issue from its timeline events.

    If the closer is a bot (e.g. pytorchmergebot), finds the linked
    PR and returns the PR author instead. Bots are never returned.
    """
    closer = None
    for event in reversed(events):
        if not isinstance(event, dict):
            continue
        if event.get("event") == "closed":
            actor = event.get("actor", {})
            closer = actor.get("login") if actor else None
            break

    if closer and closer not in BOTS:
        return closer

    # Bot closed it — find the linked PR's author
    pr_numbers = extract_pr_numbers_from_timeline(events)
    for pr_num in pr_numbers:
        try:
            pr_data = gh_json(
                [
                    "pr",
                    "view",
                    str(pr_num),
                    "--repo",
                    f"{owner}/{repo}",
                    "--json",
                    "author",
                ]
            )
        except RuntimeError:
            continue
        login = pr_data.get("author", {}).get("login")
        if login and login not in BOTS:
            return login

    return "unknown"


def extract_issue_urls_from_sheet(sheet) -> list[str]:
    """Extract all GitHub issue URLs from any cell in the sheet."""
    urls = []
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                parsed = parse_issue_url(cell.value)
                if parsed:
                    urls.append(cell.value)
            # Also check hyperlinks
            if cell.hyperlink and cell.hyperlink.target:
                parsed = parse_issue_url(cell.hyperlink.target)
                if parsed and cell.hyperlink.target not in urls:
                    urls.append(cell.hyperlink.target)
    return urls


def extract_landed_by_assignee(sheet) -> dict[str, list[str]]:
    """Extract issues marked as landed, grouped by assignee.

    Column A = assignee, column E = status. A row counts as landed
    when the status cell (case-insensitive) equals "landed".
    """
    landed: dict[str, list[str]] = defaultdict(list)
    for row in sheet.iter_rows(min_row=2):  # skip header
        assignee_cell = row[0]  # column A
        status_cell = row[4] if len(row) > 4 else None  # column E
        pr_cell = row[5] if len(row) > 5 else None  # column F
        if not status_cell or not status_cell.value:
            continue
        status = str(status_cell.value).strip().lower()
        has_pr = bool(pr_cell and pr_cell.value and str(pr_cell.value).strip())
        if status == "landed":
            pass
        elif status != "investigating" and has_pr:
            pass
        else:
            continue
        assignee = (
            str(assignee_cell.value).strip() if assignee_cell.value else "unknown"
        )
        # Find an issue URL in the row for labeling
        issue_label = None
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                parsed = parse_issue_url(cell.value)
                if parsed:
                    issue_label = f"#{parsed[2]}"
                    break
            if cell.hyperlink and cell.hyperlink.target:
                parsed = parse_issue_url(cell.hyperlink.target)
                if parsed:
                    issue_label = f"#{parsed[2]}"
                    break
        if issue_label is None:
            continue
        landed[assignee].append(issue_label)
    return landed


def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("excel_file", help="Path to Excel file (.xlsx)")
    parser.add_argument(
        "--skip-sheets",
        nargs="*",
        default=["priority_issues_20260304"],
        help="Sheet names to skip (default: priority_issues_20260304)",
    )
    args = parser.parse_args()

    wb = openpyxl.load_workbook(args.excel_file, data_only=True)

    total_closed = 0
    closed_per_area: dict[str, list[str]] = defaultdict(list)
    closed_by_user: dict[str, list[str]] = defaultdict(list)
    highpri_closed_by_user: dict[str, list[str]] = defaultdict(list)
    reviews_by_user: dict[str, list[str]] = defaultdict(list)
    landed_by_assignee: dict[str, list[str]] = defaultdict(list)
    landed_per_area: dict[str, list[str]] = defaultdict(list)

    seen_issues: set[tuple[str, str, int]] = set()
    skip = set(args.skip_sheets or [])

    for sheet_name in wb.sheetnames:
        if sheet_name in skip:
            print(f"\n--- Skipping sheet: {sheet_name} ---")
            continue
        sheet = wb[sheet_name]
        for assignee, issues in extract_landed_by_assignee(sheet).items():
            landed_by_assignee[assignee].extend(issues)
            landed_per_area[sheet_name].extend(issues)
        urls = extract_issue_urls_from_sheet(sheet)
        if not urls:
            continue

        print(f"\n--- Area: {sheet_name} ({len(urls)} issues found) ---")

        area_closed = 0
        for url in urls:
            parsed = parse_issue_url(url)
            if not parsed:
                continue
            owner, repo, number = parsed

            if (owner, repo, number) in seen_issues:
                continue
            seen_issues.add((owner, repo, number))

            try:
                info = get_issue_info(owner, repo, number)
            except RuntimeError as e:
                print(f"  Warning: could not fetch {owner}/{repo}#{number}: {e}")
                continue

            state = info.get("state", "").upper()
            title = info.get("title", "")
            print(f"  #{number}: {title} [{state}]")

            if state != "CLOSED":
                continue

            issue_label = f"#{number}: {title}"
            total_closed += 1
            area_closed += 1
            closed_per_area[sheet_name].append(issue_label)

            issue_labels = {
                lbl.get("name", "").lower() for lbl in info.get("labels", [])
            }
            is_high_pri = bool(issue_labels & HIGH_PRIORITY_LABELS)

            # Fetch timeline once, reuse for closer + reviewer analysis
            timeline = get_issue_timeline(owner, repo, number)

            # 1) If closed by a PR (via "Fixes #N"), attribute to PR author
            # 2) Otherwise, find the closer from the timeline
            closing_prs = info.get("closedByPullRequestsReferences", [])
            if closing_prs:
                for pr_ref in closing_prs:
                    user = pr_ref.get("author", {}).get("login", "unknown")
                    closed_by_user[user].append(issue_label)
                    if is_high_pri:
                        highpri_closed_by_user[user].append(issue_label)
            else:
                closer = find_closer_from_timeline(timeline, owner, repo)
                closed_by_user[closer].append(issue_label)
                if is_high_pri:
                    highpri_closed_by_user[closer].append(issue_label)

            # Count reviewers on all PRs linked to this issue
            linked_pr_numbers = extract_pr_numbers_from_timeline(timeline)
            for pr_num in linked_pr_numbers:
                try:
                    pr_data = gh_json(
                        [
                            "pr",
                            "view",
                            str(pr_num),
                            "--repo",
                            f"{owner}/{repo}",
                            "--json",
                            "number,reviews",
                        ]
                    )
                except RuntimeError:
                    continue
                seen_reviewers = set()
                for review in pr_data.get("reviews", []):
                    reviewer = review.get("author", {}).get("login", "unknown")
                    if reviewer not in seen_reviewers:
                        seen_reviewers.add(reviewer)
                        reviews_by_user[reviewer].append(
                            f"PR #{pr_num} (for {issue_label})"
                        )

        print(f"  Closed in this area: {area_closed}")

    # Print summary
    print("\n" + "=" * 60)
    print("SUMMARY")
    print("=" * 60)

    def print_grouped(header: str, data: dict[str, list[str]]) -> None:
        print(f"\n{header}")
        for key, items in sorted(data.items(), key=lambda x: -len(x[1])):
            print(f"   {key}: {len(items)}")
            for item in items:
                print(f"     - {item}")

    print(f"\n1. Total closed issues: {total_closed}")
    print_grouped("2. Closed issues per area:", closed_per_area)
    print_grouped("3. Issues closed by user:", closed_by_user)
    print_grouped(
        "4. High priority / UBN issues closed by user:", highpri_closed_by_user
    )
    print_grouped("5. PR reviews by user (top reviewers):", reviews_by_user)
    print_grouped("6. Landed issues by assignee (from sheet):", landed_by_assignee)
    print_grouped("7. Landed issues per area (from sheet):", landed_per_area)


if __name__ == "__main__":
    main()
